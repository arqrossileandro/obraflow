#!/usr/bin/env python3
"""
Reinserta Serena Folk con estructura correcta:
- 13 módulos como raíz
- Cada módulo tiene subtareas: Plomería MXX, Albañilería MXX, Hormigón MXX
- Cada gremio tiene sus items como sub-subtareas con fechas reales del cash flow

Antes borra todo lo existente de Serena Folk.
"""
import json
import sys
import os
import uuid
import urllib.request
import urllib.error
from datetime import datetime, timedelta
import openpyxl
import calendar

SUPABASE_URL = "https://dfstasoowrezcmouwqip.supabase.co"
SUPABASE_ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRmc3Rhc29vd3JlemNtb3V3cWlwIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODMyOTUyNjcsImV4cCI6MjA5ODg3MTI2N30.NGR9-0nRh9Ir6i42z9F0op3Y_9QIx5T1wyT45Fgq5V8"

MONTHS_ES = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
    'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12,
    'SEPTIEMBRE': 9, 'SETIEMBRE': 9,
}

def api(method, path, body=None, jwt=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {jwt if jwt else SUPABASE_ANON_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    data = json.dumps(body).encode() if body else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        body_text = e.read().decode()
        print(f"  ERROR {e.code} en {method} {path}: {body_text[:300]}")
        return None
    except Exception as e:
        print(f"  ERROR: {e}")
        return None

def add_days(iso_date, days):
    d = datetime.strptime(iso_date, '%Y-%m-%d')
    return (d + timedelta(days=days)).strftime('%Y-%m-%d')

def last_day_of_month(year, month):
    return calendar.monthrange(year, month)[1]

def quincena_to_dates(quincena_idx, half, start_year=2026, start_month=5):
    """
    Convierte índice de quincena (0-based) + half (0=Q1, 1=Q2) a (start_date, end_date).
    C1 = start_month, C2 = start_month+1, etc.
    """
    total_months_offset = quincena_idx
    year = start_year + (start_month - 1 + total_months_offset) // 12
    month = (start_month - 1 + total_months_offset) % 12 + 1

    if half == 0:  # Q1 = days 1-15
        start = f"{year:04d}-{month:02d}-01"
        end = f"{year:04d}-{month:02d}-15"
    else:  # Q2 = days 16-last
        start = f"{year:04d}-{month:02d}-16"
        end = f"{year:04d}-{month:02d}-{last_day_of_month(year, month):02d}"

    return start, end

# ============================================================================
# Extraer items de un sheet de módulo con sus fechas reales
# ============================================================================
def extract_module_items(ws, start_month_name):
    """
    Extrae items del sheet de un módulo con fechas calculadas desde el cash flow.
    Retorna lista de items con: name, pct, amount, finalizado, start_date, end_date
    """
    start_month = MONTHS_ES.get(start_month_name.upper(), 5) if start_month_name else 5

    # Mapear columnas a quincenas
    # Row 6 (index 5) tiene: ITEM, DESCRIPCIÓN, % ITEM, MONTO TOTAL, FINALIZADO, C1, '', C2, '', C3...
    # Row 7 (index 6) tiene: ..., Q1, Q2, Q1, Q2...
    # Las columnas de quincenas empiezan en col 5 (F)

    items = []
    current_section = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        item_num = row[0] if len(row) > 0 else None
        desc = row[1] if len(row) > 1 else None
        pct = row[2] if len(row) > 2 else None
        monto = row[3] if len(row) > 3 else None
        finalizado = row[4] if len(row) > 4 else None

        if not item_num and not desc:
            continue

        if item_num and isinstance(item_num, (int, float)):
            # Es item principal (sección)
            current_section = str(desc).strip() if desc else None
            # Para items principales con monto, calcular fechas desde las columnas de quincena
            start_date, end_date = None, None
            for col_idx in range(5, min(len(row), 25)):
                val = row[col_idx]
                if val and isinstance(val, (int, float)) and val > 0:
                    q_idx = (col_idx - 5) // 2
                    half = (col_idx - 5) % 2
                    s, e = quincena_to_dates(q_idx, half, 2026, start_month)
                    if start_date is None or s < start_date:
                        start_date = s
                    if end_date is None or e > end_date:
                        end_date = e

            if start_date is None:
                # Si no hay valores en quincenas, usar fecha genérica
                start_date = f"2026-{start_month:02d}-01"
                end_date = add_days(start_date, 90)

            items.append({
                'section': current_section,
                'name': str(desc).strip() if desc else '',
                'pct': float(pct) if pct else 0,
                'amount': float(monto) if monto else 0,
                'finalizado': float(finalizado) if finalizado else 0,
                'is_section_header': True,
                'start_date': start_date,
                'end_date': end_date,
            })
        elif desc and not item_num:
            # Es subitem
            start_date, end_date = None, None
            for col_idx in range(5, min(len(row), 25)):
                val = row[col_idx]
                if val and isinstance(val, (int, float)) and val > 0:
                    q_idx = (col_idx - 5) // 2
                    half = (col_idx - 5) % 2
                    s, e = quincena_to_dates(q_idx, half, 2026, start_month)
                    if start_date is None or s < start_date:
                        start_date = s
                    if end_date is None or e > end_date:
                        end_date = e

            if start_date is None:
                start_date = f"2026-{start_month:02d}-01"
                end_date = add_days(start_date, 30)

            items.append({
                'section': current_section,
                'name': str(desc).strip(),
                'pct': float(pct) if pct else 0,
                'amount': 0,
                'finalizado': float(finalizado) if finalizado else 0,
                'is_section_header': False,
                'start_date': start_date,
                'end_date': end_date,
            })

    return items, start_month

# ============================================================================
# Extraer llenadas de hormigón
# ============================================================================
def extract_hormigon():
    print("Extrayendo llenadas de hormigón...")
    wb = openpyxl.load_workbook('/home/z/my-project/upload/SF - PLAN INTEGRAL.xlsb.xlsx', data_only=True)
    ws = wb['LISTA DE LLENADAS']

    llenadas = []
    current_date = None
    current_period = None

    for row in ws.iter_rows(values_only=True):
        first = row[0]
        second = row[1]

        if first and str(first).upper() in MONTHS_ES:
            current_period = str(first).upper()
            continue

        if first:
            if isinstance(first, datetime):
                current_date = first.strftime('%Y-%m-%d')
                continue
            s = str(first).strip()
            if s and s[0].isdigit():
                try:
                    current_date = datetime.strptime(s[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
                    continue
                except:
                    pass

        if current_date and second and isinstance(second, str):
            desc = second.strip()
            m3 = row[2] if len(row) > 2 and row[2] else 0
            modules = []
            for cell in row[3:10]:
                if cell and isinstance(cell, str) and cell.strip().startswith('M'):
                    modules.append(cell.strip())

            if m3 and isinstance(m3, (int, float)) and m3 > 0 and 'TOTALES' not in desc.upper():
                llenadas.append({
                    'date': current_date,
                    'description': desc,
                    'm3': float(m3),
                    'modules': modules,
                })

    # Agrupar por módulo
    by_module = {}
    for l in llenadas:
        for m in l.get('modules', []):
            if m not in by_module:
                by_module[m] = []
            by_module[m].append(l)
        if not l.get('modules'):
            if 'GENERAL' not in by_module:
                by_module['GENERAL'] = []
            by_module['GENERAL'].append(l)

    print(f"  {len(llenadas)} llenadas en {len(by_module)} módulos")
    return by_module

# ============================================================================
# MAIN
# ============================================================================
def main():
    if len(sys.argv) < 2:
        print("Uso: python3 reinsert_serena_folk.py <JWT>")
        sys.exit(1)
    jwt = sys.argv[1].strip()

    # 1. Encontrar y borrar Serena Folk existente
    print("\n=== Buscando Serena Folk existente ===")
    obras = api('GET', 'obras?name=eq.Serena%20Folk&select=id,name', jwt=jwt)
    if obras and len(obras) > 0:
        obra_id_old = obras[0]['id']
        print(f"  Encontrada: {obra_id_old}")

        # Borrar todas las tareas (cascada debería borrar subtareas, deps, etc.)
        print("  Borrando tareas...")
        api('DELETE', f'tasks?obra_id=eq.{obra_id_old}', jwt=jwt)
        # Borrar materiales
        api('DELETE', f'materials?obra_id=eq.{obra_id_old}', jwt=jwt)
        # Borrar chat messages
        api('DELETE', f'chat_messages?obra_id=eq.{obra_id_old}', jwt=jwt)
        # Borrar obra
        print("  Borrando obra...")
        api('DELETE', f'obras?id=eq.{obra_id_old}', jwt=jwt)
        print("  Serena Folk vieja borrada")
    else:
        print("  No existe Serena Folk previa")

    # 2. Crear obra nueva
    print("\n=== Creando obra Serena Folk ===")
    obra_id = str(uuid.uuid4())
    obra = {
        'id': obra_id,
        'name': 'Serena Folk',
        'client': 'Serena Folk',
        'address': 'Complejo de 13 módulos — Hormigón expuesto',
        'start_date': '2026-04-01',
        'end_date': '2027-06-30',
        'budget': 250000000,
        'color': '#22c55e',
        'status': 'en_curso',
        'progress': 0,
    }
    result = api('POST', 'obras', obra, jwt)
    if not result:
        print("ERROR: No se pudo crear la obra")
        sys.exit(1)
    print(f"  Obra creada: {obra_id}")

    # 3. Extraer datos de Plomería
    print("\n=== Extrayendo Plomería Rodrigo ===")
    wb_plom = openpyxl.load_workbook('/home/z/my-project/upload/SF - PLOMERIA RODRIGO (MITAD DE MES).xlsx', data_only=True)
    plomeria_data = {}  # m_key -> {items, start_month}
    for sheet_name in wb_plom.sheetnames:
        if not (sheet_name.startswith('M') and sheet_name[1:].isdigit()):
            continue
        module_num = int(sheet_name[1:])
        m_key = f"M{module_num:02d}"
        ws = wb_plom[sheet_name]
        # Leer mes de inicio
        start_month_name = None
        for cell in ws[5]:
            if cell.value and str(cell.value).upper() in MONTHS_ES:
                start_month_name = str(cell.value).upper()
                break
        items, start_month = extract_module_items(ws, start_month_name)
        monto = float(ws['D3'].value) if ws['D3'].value else 0
        plomeria_data[m_key] = {
            'items': items,
            'start_month': start_month,
            'monto': monto,
        }
        print(f"  {m_key}: {len(items)} items, monto ${monto:,.0f}, inicio mes {start_month}")

    # 4. Extraer datos de Albañilería
    print("\n=== Extrayendo Albañilería Beltrán ===")
    wb_alb = openpyxl.load_workbook('/home/z/my-project/upload/SF - ALBAÑILERIA BELTRAN (MITAD DE MES).xlsx', data_only=True)
    albanileria_data = {}
    for sheet_name in wb_alb.sheetnames:
        if not (sheet_name.startswith('M') and sheet_name[1:].isdigit()):
            continue
        module_num = int(sheet_name[1:])
        m_key = f"M{module_num:02d}"
        ws = wb_alb[sheet_name]
        start_month_name = None
        for cell in ws[5]:
            if cell.value and str(cell.value).upper() in MONTHS_ES:
                start_month_name = str(cell.value).upper()
                break
        items, start_month = extract_module_items(ws, start_month_name)
        monto = float(ws['C3'].value) if ws['C3'].value else 0
        albanileria_data[m_key] = {
            'items': items,
            'start_month': start_month,
            'monto': monto,
        }
        print(f"  {m_key}: {len(items)} items, monto ${monto:,.0f}, inicio mes {start_month}")

    # 5. Extraer llenadas de hormigón
    print("\n=== Extrayendo Hormigón ===")
    hormigon_by_module = extract_hormigon()

    # 6. Crear 13 módulos como raíz
    print("\n=== Creando 13 módulos raíz ===")
    MODULO_COLORS = ['#10b981', '#14b8a6', '#06b6d4', '#0ea5e9', '#3b82f6', '#6366f1', '#8b5cf6', '#a855f7', '#ec4899', '#f43f5e', '#ef4444', '#f97316', '#eab308']

    modulo_ids = {}
    modulo_dates = {}  # m_key -> {min_start, max_end}

    for i in range(1, 14):
        m_key = f"M{i:02d}"
        m_id = str(uuid.uuid4())
        modulo_ids[m_key] = m_id

        # Calcular fechas del módulo = min/max de todos sus items
        all_starts = []
        all_ends = []
        if m_key in plomeria_data:
            for item in plomeria_data[m_key]['items']:
                if item.get('start_date'): all_starts.append(item['start_date'])
                if item.get('end_date'): all_ends.append(item['end_date'])
        if m_key in albanileria_data:
            for item in albanileria_data[m_key]['items']:
                if item.get('start_date'): all_starts.append(item['start_date'])
                if item.get('end_date'): all_ends.append(item['end_date'])
        if m_key in hormigon_by_module:
            for l in hormigon_by_module[m_key]:
                all_starts.append(l['date'])
                all_ends.append(l['date'])

        min_start = min(all_starts) if all_starts else '2026-05-01'
        max_end = max(all_ends) if all_ends else '2026-12-31'
        modulo_dates[m_key] = {'start': min_start, 'end': max_end}

        # Calcular progreso basado en items finalizados
        total_items = 0
        completed_items = 0
        if m_key in plomeria_data:
            for item in plomeria_data[m_key]['items']:
                if item.get('is_section_header'):
                    total_items += 1
                    if item.get('finalizado', 0) >= 1:
                        completed_items += 1
        if m_key in albanileria_data:
            for item in albanileria_data[m_key]['items']:
                if item.get('is_section_header'):
                    total_items += 1
                    if item.get('finalizado', 0) >= 1:
                        completed_items += 1

        progress = int((completed_items / total_items) * 100) if total_items > 0 else 0

        task = {
            'id': m_id,
            'obra_id': obra_id,
            'parent_id': None,
            'name': f'Módulo {i:02d}',
            'description': f'Módulo {i:02d} del complejo Serena Folk',
            'start_date': min_start,
            'end_date': max_end,
            'progress': progress,
            'progress_mode': 'manual',
            'manual_progress': progress,
            'guild': 'Estructura',
            'color': MODULO_COLORS[i-1],
            'priority': 'alta',
            'status': 'en_curso' if progress > 0 else 'no_iniciada',
        }
        api('POST', 'tasks', task, jwt)
        print(f"  M{i:02d} creado: {min_start} → {max_end} ({progress}%)")

    # 7. Crear subtareas de gremio bajo cada módulo
    COLOR_PLOMERIA = '#0ea5e9'
    COLOR_ALBANILERIA = '#f97316'
    COLOR_HORMIGON = '#64748b'

    # 7a. Plomería por módulo
    print("\n=== Creando Plomería por módulo ===")
    for m_key, mod_data in plomeria_data.items():
        if m_key not in modulo_ids:
            continue
        m_id = modulo_ids[m_key]

        # Fechas de la plomería de este módulo
        starts = [item['start_date'] for item in mod_data['items'] if item.get('start_date')]
        ends = [item['end_date'] for item in mod_data['items'] if item.get('end_date')]
        min_start = min(starts) if starts else '2026-05-01'
        max_end = max(ends) if ends else '2026-08-31'

        # Progreso
        total = sum(1 for i in mod_data['items'] if i.get('is_section_header'))
        done = sum(1 for i in mod_data['items'] if i.get('is_section_header') and i.get('finalizado', 0) >= 1)
        progress = int((done / total) * 100) if total > 0 else 0

        # Tarea "Plomería MXX"
        plom_id = str(uuid.uuid4())
        plom_task = {
            'id': plom_id,
            'obra_id': obra_id,
            'parent_id': m_id,
            'name': f'Plomería {m_key}',
            'description': f'Plomería Rodrigo del {m_key} — monto: ${mod_data["monto"]:,.0f}',
            'start_date': min_start,
            'end_date': max_end,
            'progress': progress,
            'progress_mode': 'manual',
            'manual_progress': progress,
            'guild': 'Plomería',
            'color': COLOR_PLOMERIA,
            'priority': 'media',
            'status': 'finalizada' if progress >= 100 else ('en_curso' if progress > 0 else 'no_iniciada'),
            'labor_cost': mod_data['monto'],
        }
        api('POST', 'tasks', plom_task, jwt)

        # Sub-subtareas: items
        for item in mod_data['items']:
            if not item.get('is_section_header'):
                continue  # Solo crear tareas para items principales con monto
            if item.get('amount', 0) == 0:
                continue
            item_id = str(uuid.uuid4())
            item_progress = int(item.get('finalizado', 0) * 100)
            item_task = {
                'id': item_id,
                'obra_id': obra_id,
                'parent_id': plom_id,
                'name': item['name'],
                'description': f'{item["name"]} — {item["pct"]*100:.1f}% del módulo',
                'start_date': item['start_date'],
                'end_date': item['end_date'],
                'progress': item_progress,
                'progress_mode': 'manual',
                'manual_progress': item_progress,
                'guild': 'Plomería',
                'color': COLOR_PLOMERIA,
                'priority': 'baja',
                'status': 'finalizada' if item_progress >= 100 else ('en_curso' if item_progress > 0 else 'no_iniciada'),
                'labor_cost': item['amount'],
                'repercussion_percent': item['pct'] * 100,
            }
            api('POST', 'tasks', item_task, jwt)
        print(f"  Plomería {m_key}: {min_start} → {max_end} ({progress}%)")

    # 7b. Albañilería por módulo
    print("\n=== Creando Albañilería por módulo ===")
    for m_key, mod_data in albanileria_data.items():
        if m_key not in modulo_ids:
            continue
        m_id = modulo_ids[m_key]

        starts = [item['start_date'] for item in mod_data['items'] if item.get('start_date')]
        ends = [item['end_date'] for item in mod_data['items'] if item.get('end_date')]
        min_start = min(starts) if starts else '2026-05-01'
        max_end = max(ends) if ends else '2026-12-31'

        total = sum(1 for i in mod_data['items'] if i.get('is_section_header'))
        done = sum(1 for i in mod_data['items'] if i.get('is_section_header') and i.get('finalizado', 0) >= 1)
        progress = int((done / total) * 100) if total > 0 else 0

        alb_id = str(uuid.uuid4())
        alb_task = {
            'id': alb_id,
            'obra_id': obra_id,
            'parent_id': m_id,
            'name': f'Albañilería {m_key}',
            'description': f'Albañilería Beltrán del {m_key} — monto: ${mod_data["monto"]:,.0f}',
            'start_date': min_start,
            'end_date': max_end,
            'progress': progress,
            'progress_mode': 'manual',
            'manual_progress': progress,
            'guild': 'Albañilería',
            'color': COLOR_ALBANILERIA,
            'priority': 'alta',
            'status': 'finalizada' if progress >= 100 else ('en_curso' if progress > 0 else 'no_iniciada'),
            'labor_cost': mod_data['monto'],
        }
        api('POST', 'tasks', alb_task, jwt)

        for item in mod_data['items']:
            if not item.get('is_section_header'):
                continue
            if item.get('amount', 0) == 0:
                continue
            item_id = str(uuid.uuid4())
            item_progress = int(item.get('finalizado', 0) * 100)
            item_task = {
                'id': item_id,
                'obra_id': obra_id,
                'parent_id': alb_id,
                'name': item['name'],
                'description': f'{item["name"]} — {item["pct"]*100:.1f}%',
                'start_date': item['start_date'],
                'end_date': item['end_date'],
                'progress': item_progress,
                'progress_mode': 'manual',
                'manual_progress': item_progress,
                'guild': 'Albañilería',
                'color': COLOR_ALBANILERIA,
                'priority': 'media',
                'status': 'finalizada' if item_progress >= 100 else ('en_curso' if item_progress > 0 else 'no_iniciada'),
                'labor_cost': item['amount'],
                'repercussion_percent': item['pct'] * 100,
            }
            api('POST', 'tasks', item_task, jwt)
        print(f"  Albañilería {m_key}: {min_start} → {max_end} ({progress}%)")

    # 7c. Hormigón por módulo
    print("\n=== Creando Hormigón por módulo ===")
    for m_key, llenadas in hormigon_by_module.items():
        if m_key == 'GENERAL':
            continue  # Las llenadas generales las asignamos al módulo más cercano o las saltamos
        if m_key not in modulo_ids:
            # Normalizar M6 -> M06
            m_key_norm = f"M{int(m_key[1:]):02d}"
            if m_key_norm in modulo_ids:
                m_key = m_key_norm
            else:
                continue
        m_id = modulo_ids[m_key]

        if not llenadas:
            continue

        starts = [l['date'] for l in llenadas]
        ends = [l['date'] for l in llenadas]
        min_start = min(starts)
        max_end = max(ends)
        total_m3 = sum(l['m3'] for l in llenadas)

        # Progreso: llenadas pasadas = finalizadas
        today = datetime.now()
        done = sum(1 for l in llenadas if datetime.strptime(l['date'], '%Y-%m-%d') < today)
        progress = int((done / len(llenadas)) * 100) if llenadas else 0

        horm_id = str(uuid.uuid4())
        horm_task = {
            'id': horm_id,
            'obra_id': obra_id,
            'parent_id': m_id,
            'name': f'Hormigón {m_key}',
            'description': f'Llenadas del {m_key} — {len(llenadas)} tiradas, {total_m3:.1f} m³',
            'start_date': min_start,
            'end_date': max_end,
            'progress': progress,
            'progress_mode': 'manual',
            'manual_progress': progress,
            'guild': 'Hormigón',
            'color': COLOR_HORMIGON,
            'priority': 'alta',
            'status': 'finalizada' if progress >= 100 else ('en_curso' if progress > 0 else 'no_iniciada'),
            'materials_cost': int(total_m3 * 80000),
        }
        api('POST', 'tasks', horm_task, jwt)

        # Sub-subtareas: cada llenada
        for l in llenadas:
            l_id = str(uuid.uuid4())
            l_date = datetime.strptime(l['date'], '%Y-%m-%d')
            if l_date < today:
                l_status = 'finalizada'
                l_progress = 100
            elif (l_date - today).days <= 30:
                l_status = 'en_curso'
                l_progress = 50
            else:
                l_status = 'no_iniciada'
                l_progress = 0

            item_task = {
                'id': l_id,
                'obra_id': obra_id,
                'parent_id': horm_id,
                'name': f'{l["description"]} ({l["m3"]:.1f} m³)',
                'description': f'{l["description"]} — {l["m3"]} m³ — fecha: {l["date"]}',
                'start_date': l['date'],
                'end_date': add_days(l['date'], 1),
                'progress': l_progress,
                'progress_mode': 'manual',
                'manual_progress': l_progress,
                'guild': 'Hormigón',
                'color': COLOR_HORMIGON,
                'priority': 'alta' if l_status == 'en_curso' else 'media',
                'status': l_status,
                'materials_cost': int(l['m3'] * 80000),
            }
            api('POST', 'tasks', item_task, jwt)
        print(f"  Hormigón {m_key}: {min_start} → {max_end} ({progress}%, {total_m3:.1f} m³)")

    # 8. Crear tarea raíz para llenadas generales (sin módulo asignado)
    if 'GENERAL' in hormigon_by_module and hormigon_by_module['GENERAL']:
        print("\n=== Creando Hormigón General ===")
        gen_llenadas = hormigon_by_module['GENERAL']
        starts = [l['date'] for l in gen_llenadas]
        ends = [l['date'] for l in gen_llenadas]
        min_start = min(starts)
        max_end = max(ends)
        total_m3 = sum(l['m3'] for l in gen_llenadas)

        gen_id = str(uuid.uuid4())
        gen_task = {
            'id': gen_id,
            'obra_id': obra_id,
            'parent_id': None,
            'name': 'Hormigón General',
            'description': f'Llenadas sin módulo específico — {len(gen_llenadas)} tiradas, {total_m3:.1f} m³',
            'start_date': min_start,
            'end_date': max_end,
            'progress': 50,
            'progress_mode': 'manual',
            'manual_progress': 50,
            'guild': 'Hormigón',
            'color': COLOR_HORMIGON,
            'priority': 'media',
            'status': 'en_curso',
        }
        api('POST', 'tasks', gen_task, jwt)
        for l in gen_llenadas:
            l_id = str(uuid.uuid4())
            item_task = {
                'id': l_id,
                'obra_id': obra_id,
                'parent_id': gen_id,
                'name': f'{l["description"]} ({l["m3"]:.1f} m³)',
                'description': f'{l["description"]} — {l["m3"]} m³ — {l["date"]}',
                'start_date': l['date'],
                'end_date': add_days(l['date'], 1),
                'progress': 100,
                'progress_mode': 'manual',
                'manual_progress': 100,
                'guild': 'Hormigón',
                'color': COLOR_HORMIGON,
                'priority': 'baja',
                'status': 'finalizada',
                'materials_cost': int(l['m3'] * 80000),
            }
            api('POST', 'tasks', item_task, jwt)
        print(f"  Hormigón General: {len(gen_llenadas)} llenadas, {total_m3:.1f} m³")

    print("\n=== ¡LISTO! ===")
    print("Serena Folk recreada con estructura correcta:")
    print("  - 13 módulos como raíz (M01-M13)")
    print("  - Cada módulo tiene: Plomería MXX + Albañilería MXX + Hormigón MXX")
    print("  - Cada gremio tiene sus items con fechas reales del cash flow")
    print("  - Las barras de los padres duran desde la primera hasta la última subtarea")

if __name__ == '__main__':
    main()
