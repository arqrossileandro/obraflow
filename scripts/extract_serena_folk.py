#!/usr/bin/env python3
"""
Genera un archivo SQL para insertar la obra Serena Folk con todos sus datos:
- Obra + miembros
- 13 módulos (tareas raíz) — M01 a M13
- 2 gremios con cash flow: Albañilería Beltrán y Plomería Rodrigo
- Plan integral de llenadas de hormigón

Output: /home/z/my-project/download/serena-folk-seed.sql
"""
import openpyxl
import json
import uuid
from datetime import datetime, timedelta
import os
import urllib.request
import urllib.error

# Configuración Supabase
SUPABASE_URL = "https://dfstasoowrezcmouwqip.supabase.co"
# Usar la service role key que permite insertar con RLS bypass
# Como no la tenemos, usaremos inserts directos via API REST con el anon key
# y el usuario ya autenticado será el creador
SUPABASE_ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRmc3Rhc29vd3JlemNtb3V3cWlwIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODMyOTUyNjcsImV4cCI6MjA5ODg3MTI2N30.NGR9-0nRh9Ir6i42z9F0op3Y_9QIx5T1wyT45Fgq5V8"

# Directorios de los archivos Excel
ALBANILERIA_FILE = '/home/z/my-project/upload/SF - ALBAÑILERIA BELTRAN (MITAD DE MES).xlsx'
PLOMERIA_FILE = '/home/z/my-project/upload/SF - PLOMERIA RODRIGO (MITAD DE MES).xlsx'
PLAN_INTEGRAL_FILE = '/home/z/my-project/upload/SF - PLAN INTEGRAL.xlsb.xlsx'

# Mapeo de meses en español → número
MONTH_MAP = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
    'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12,
    'SEPTIEMBRE': 9, 'SETIEMBRE': 9,
}

def parse_month_year(s):
    """Convertir 'MAYO' o 'MAYO 2026' a '2026-05-01'."""
    if not s:
        return None
    parts = str(s).upper().strip().split()
    if not parts:
        return None
    month_name = parts[0]
    if month_name not in MONTH_MAP:
        return None
    month = MONTH_MAP[month_name]
    year = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 2026
    return f"{year:04d}-{month:02d}-01"

def parse_date(s):
    """Convertir '2026-04-07 00:00:00' o similar a '2026-04-07'."""
    if not s:
        return None
    if isinstance(s, datetime):
        return s.strftime('%Y-%m-%d')
    s = str(s).strip()
    if not s:
        return None
    # Intentar parsear YYYY-MM-DD
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
    except:
        return None

# ============================================================================
# Extraer datos de PLOMERIA
# ============================================================================
def extract_plomeria():
    """Extrae tareas de plomería del archivo."""
    print("Extrayendo Plomería Rodrigo...")
    wb = openpyxl.load_workbook(PLOMERIA_FILE, data_only=True)

    # Leer RESUMEN para montos totales
    ws = wb['RESUMEN']
    monto_total = None
    for row in ws.iter_rows(values_only=True):
        if row[0] and 'PLOMERIA TOTAL' in str(row[0]):
            monto_total = row[3]
            break
    print(f"  Monto total Plomería: ${monto_total:,.0f}" if monto_total else "  No se encontró monto")

    # Leer cada módulo (M01-M13)
    tasks_by_module = {}
    for sheet_name in wb.sheetnames:
        if not (sheet_name.startswith('M') and sheet_name[1:].isdigit()):
            continue
        module_num = int(sheet_name[1:])
        module_key = f"M{module_num:02d}"
        ws = wb[sheet_name]

        # Leer mes de inicio (fila 5 tiene los meses)
        start_month = None
        for cell in ws[5]:
            if cell.value and str(cell.value).upper() in MONTH_MAP:
                start_month = str(cell.value).upper()
                break

        # Extraer items del cash flow
        items = []
        current_section = None
        for row in ws.iter_rows(min_row=6, values_only=True):
            item_num = row[0]
            desc = row[1]
            pct = row[2]
            monto = row[3]
            finalizado = row[4]

            if not item_num and not desc:
                continue
            if item_num and isinstance(item_num, (int, float)):
                # Es item principal
                current_section = str(desc).strip() if desc else None
                items.append({
                    'section': current_section,
                    'name': str(desc).strip() if desc else '',
                    'pct': float(pct) if pct else 0,
                    'amount': float(monto) if monto else 0,
                    'finalizado': float(finalizado) if finalizado else 0,
                    'is_section_header': True,
                })
            elif desc and not item_num:
                # Es subitem
                items.append({
                    'section': current_section,
                    'name': str(desc).strip(),
                    'pct': float(pct) if pct else 0,
                    'amount': 0,  # subitems no tienen monto
                    'finalizado': float(finalizado) if finalizado else 0,
                    'is_section_header': False,
                })

        tasks_by_module[module_key] = {
            'start_month': start_month,
            'items': items,
            'monto_total': float(ws['D3'].value) if ws['D3'].value else 0,
        }
        print(f"  {module_key}: {len(items)} items, monto módulo ${ws['D3'].value:,.0f}" if ws['D3'].value else f"  {module_key}: {len(items)} items")

    return {
        'monto_total': monto_total or 166000000,
        'modules': tasks_by_module,
    }

# ============================================================================
# Extraer datos de ALBAÑILERIA
# ============================================================================
def extract_albanileria():
    """Extrae tareas de albañilería del archivo."""
    print("\nExtrayendo Albañilería Beltrán...")
    wb = openpyxl.load_workbook(ALBANILERIA_FILE, data_only=True)

    # Leer CASH FLOW PROYECTADO para montos
    ws = wb['CASH FLOW PROYECTADO']
    monto_obra = ws['C3'].value
    print(f"  Monto obra inicial: ${monto_obra:,.0f}" if monto_obra else "  No se encontró monto")

    # Leer pagos quincenales
    ws_pagos = wb['PAGOS QUINCENALES']
    pagos = []
    current_period = None
    for row in ws_pagos.iter_rows(values_only=True):
        first_cell = row[0]
        if first_cell and str(first_cell).upper() in MONTH_MAP:
            current_period = str(first_cell).upper()
        elif first_cell and isinstance(first_cell, (int, float)) and current_period:
            # Es un pago Q1 o Q2
            pagos.append({
                'period': current_period,
                'amount': float(first_cell),
            })

    print(f"  Pagos quincenales encontrados: {len(pagos)}")

    # Leer cada módulo
    tasks_by_module = {}
    for sheet_name in wb.sheetnames:
        if not (sheet_name.startswith('M') and sheet_name[1:].isdigit()):
            continue
        module_num = int(sheet_name[1:])
        module_key = f"M{module_num:02d}"
        ws = wb[sheet_name]

        # Mes de inicio
        start_month = None
        for cell in ws[5]:
            if cell.value and str(cell.value).upper() in MONTH_MAP:
                start_month = str(cell.value).upper()
                break

        # Items
        items = []
        current_section = None
        for row in ws.iter_rows(min_row=6, values_only=True):
            item_num = row[0]
            desc = row[1]
            pct = row[2]
            monto = row[3]
            finalizado = row[4]

            if not item_num and not desc:
                continue
            if item_num and isinstance(item_num, (int, float)):
                current_section = str(desc).strip() if desc else None
                items.append({
                    'section': current_section,
                    'name': str(desc).strip() if desc else '',
                    'pct': float(pct) if pct else 0,
                    'amount': float(monto) if monto else 0,
                    'finalizado': float(finalizado) if finalizado else 0,
                    'is_section_header': True,
                })
            elif desc and not item_num:
                items.append({
                    'section': current_section,
                    'name': str(desc).strip(),
                    'pct': float(pct) if pct else 0,
                    'amount': 0,
                    'finalizado': float(finalizado) if finalizado else 0,
                    'is_section_header': False,
                })

        tasks_by_module[module_key] = {
            'start_month': start_month,
            'items': items,
            'monto_total': float(ws['C3'].value) if ws['C3'].value else 0,
            'anticipo': float(ws['C4'].value) if ws['C4'].value else 0,
        }
        print(f"  {module_key}: {len(items)} items, monto ${ws['C3'].value:,.0f}" if ws['C3'].value else f"  {module_key}: {len(items)} items")

    return {
        'monto_total': monto_obra or 35000000,
        'pagos_quincenales': pagos,
        'modules': tasks_by_module,
    }

# ============================================================================
# Extraer PLAN INTEGRAL (llenadas de hormigón)
# ============================================================================
def extract_plan_integral():
    """Extrae las llenadas de hormigón del plan integral."""
    print("\nExtrayendo Plan Integral (llenadas de hormigón)...")
    wb = openpyxl.load_workbook(PLAN_INTEGRAL_FILE, data_only=True)
    ws = wb['LISTA DE LLENADAS']

    llenadas = []
    current_date = None
    current_period = None  # ABRIL, MAYO, etc
    in_tirada = False

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        first = row[0]
        second = row[1]

        # Detectar cambio de mes
        if first and str(first).upper() in MONTH_MAP:
            current_period = str(first).upper()
            continue

        # Detectar fecha
        if first:
            d = parse_date(first)
            if d:
                current_date = d
                in_tirada = 'TIRADO' in str(second).upper() if second else False
                continue
            if 'TIRADO' in str(first).upper():
                in_tirado = True
                continue
            if 'INC' in str(first).upper():
                # Es una fila de item con código INC
                pass

        # Si estamos en una sección de llenada y hay descripción + m³
        if current_date and second and isinstance(second, str):
            desc = second.strip()
            m3 = row[2] if row[2] else 0

            # Buscar módulos en esta fila
            modules = []
            for cell in row[3:10]:
                if cell and isinstance(cell, str) and cell.startswith('M'):
                    modules.append(cell.strip())

            if m3 and isinstance(m3, (int, float)) and m3 > 0 and 'TOTALES' not in desc.upper():
                llenadas.append({
                    'date': current_date,
                    'description': desc,
                    'm3': float(m3),
                    'modules': modules,
                    'period': current_period,
                })

    # Agrupar por fecha
    by_date = {}
    for l in llenadas:
        d = l['date']
        if d not in by_date:
            by_date[d] = []
        by_date[d].append(l)

    print(f"  Total llenadas: {len(llenadas)}")
    print(f"  Fechas únicas: {len(by_date)}")
    for d, items in sorted(by_date.items()):
        total_m3 = sum(i['m3'] for i in items)
        mods = set()
        for i in items:
            mods.update(i['modules'])
        print(f"    {d}: {total_m3} m³ en {len(items)} elementos — módulos: {sorted(mods)}")

    return {'llenadas': llenadas, 'by_date': by_date}

# ============================================================================
# MAIN
# ============================================================================
if __name__ == '__main__':
    plomeria = extract_plomeria()
    albanileria = extract_albanileria()
    hormigon = extract_plan_integral()

    print("\n" + "="*80)
    print("RESUMEN")
    print("="*80)
    print(f"Plomería: monto total ${plomeria['monto_total']:,.0f}")
    print(f"  Módulos con datos: {len(plomeria['modules'])}")
    print(f"Albañilería: monto total ${albanileria['monto_total']:,.0f}")
    print(f"  Módulos con datos: {len(albanileria['modules'])}")
    print(f"  Pagos quincenales: {len(albanileria['pagos_quincenales'])}")
    print(f"Hormigón: {len(hormigon['llenadas'])} llenadas en {len(hormigon['by_date'])} fechas")

    # Guardar JSON intermedio
    output = {
        'plomeria': plomeria,
        'albanileria': albanileria,
        'hormigon': {
            'llenadas': hormigon['llenadas'],
        },
    }
    os.makedirs('/home/z/my-project/download', exist_ok=True)
    with open('/home/z/my-project/download/serena-folk-data.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nDatos guardados en /home/z/my-project/download/serena-folk-data.json")
