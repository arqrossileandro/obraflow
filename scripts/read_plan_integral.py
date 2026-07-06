#!/usr/bin/env python3
"""Read the PLAN INTEGRAL file in detail."""
import openpyxl

path = '/home/z/my-project/upload/SF - PLAN INTEGRAL.xlsb.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
print(f"Sheets: {wb.sheetnames}")
print()

# Focus on the key sheets
for sheet_name in ['LISTA DE LLENADAS', 'COMPLETO 13 MODULOS', 'RESUMEN ALBAÑILERIA']:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: '{sheet_name}' (rows: {ws.max_row}, cols: {ws.max_column})")
    print(f"{'='*80}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 30:
            print(f"... ({ws.max_row - 30} more rows)")
            break
        cleaned = []
        for j, v in enumerate(row):
            if j >= 12:
                break
            if v is None:
                cleaned.append('')
            elif isinstance(v, (int, float)):
                cleaned.append(v)
            else:
                s = str(v).strip()
                cleaned.append(s if s else '')
        # Skip fully empty rows
        if any(c != '' for c in cleaned):
            print(f"R{i+1}: {cleaned}")
