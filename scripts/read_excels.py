#!/usr/bin/env python3
"""Read Excel files and dump structure to understand them."""
import openpyxl
import sys
import json

def read_xlsx(path, max_rows=15, max_cols=20):
    """Read an xlsx file and print all sheets with their data."""
    print(f"\n{'='*80}")
    print(f"FILE: {path}")
    print(f"{'='*80}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"Error opening: {e}")
        return
    print(f"Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: '{sheet_name}' (dims: {ws.dimensions}, rows: {ws.max_row}, cols: {ws.max_column}) ---")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                print(f"... ({ws.max_row - max_rows} more rows)")
                break
            # Print non-empty cells only
            cleaned = []
            for j, v in enumerate(row):
                if j >= max_cols:
                    break
                if v is None:
                    cleaned.append('')
                elif isinstance(v, (int, float)):
                    cleaned.append(v)
                else:
                    s = str(v).strip()
                    cleaned.append(s if s else '')
            # Print row with row index
            print(f"R{i+1}: {cleaned}")

if __name__ == '__main__':
    files = [
        '/home/z/my-project/upload/SF - ALBAÑILERIA BELTRAN (MITAD DE MES).xlsx',
        '/home/z/my-project/upload/SF - PLOMERIA RODRIGO (MITAD DE MES).xlsx',
        '/home/z/my-project/upload/SF - PLAN INTEGRAL.xlsb.xlsx',
    ]
    for f in files:
        read_xlsx(f)
